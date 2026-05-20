import sys
import openpyxl

try:
    sys.stdout.reconfigure(encoding='utf-8')
except Exception:
    pass

wb = openpyxl.load_workbook("S.d.t.ch.c.xlsx")
sheet = wb.active

managers = []
supervisors = []
leaders = []

current_mgr = None
current_sv = None

# Let's collect all B, C, D names
# Row 5 is header, skip it for names.
# Wait, let's write a loop to extract them in order of occurrence.
all_names = [] # list of (role, name, parent_name)

for r in range(4, 68):
    if r == 5:
        continue # skip row 5 for B, C, D names
        
    b = sheet.cell(r, 2).value
    c = sheet.cell(r, 3).value
    d = sheet.cell(r, 4).value
    
    if b and str(b).strip():
        name = str(b).strip()
        all_names.append(('manager', name, None))
        current_mgr = name
        current_sv = None
        
    if c and str(c).strip():
        name = str(c).strip()
        all_names.append(('supervisor', name, current_mgr))
        current_sv = name
        
    if d and str(d).strip():
        name = str(d).strip()
        all_names.append(('leader', name, current_sv))

print(f"Total names extracted: {len(all_names)}")

# Now let's extract all usernames from Column F
usernames = []
for r in range(4, 68):
    f = sheet.cell(r, 6).value
    g = sheet.cell(r, 7).value
    if f:
        usernames.append((str(f).strip(), g))

print(f"Total usernames extracted: {len(usernames)}")

# Let's print the alignment
for idx, (role, name, parent) in enumerate(all_names):
    username = usernames[idx][0] if idx < len(usernames) else "MISSING"
    pass_val = usernames[idx][1] if idx < len(usernames) else "MISSING"
    print(f"{idx+1:02d}: Role={role:<10} Name={name:<25} Parent={str(parent):<20} Username={username:<15} Pass={pass_val}")

