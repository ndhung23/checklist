import sys
import openpyxl

try:
    sys.stdout.reconfigure(encoding='utf-8')
except Exception:
    pass

wb = openpyxl.load_workbook("S.d.t.ch.c.xlsx")
sheet = wb.active

users = []
curr_mgr = None
curr_sv = None
curr_tl = None

# Row 4 is first row of data:
# Row 4: Column B is None, C is None, D is None, E is None, F is 'hienMGR', G is 1
# Row 5: Column A is 'Role', B is 'manager(MGR)', C is 'Supervisor(SV)', D is 'team leader(TL:Tổ trưởng)', E is 'shift Leader(SL:Tổ phó)', F is 'maiSV1', G is 1
# Wait! Let's print rows 4 to 67 in detail.

print("--- Sequential Rows ---")
for r in range(4, 68):
    b = sheet.cell(r, 2).value
    c = sheet.cell(r, 3).value
    d = sheet.cell(r, 4).value
    e = sheet.cell(r, 5).value
    f = sheet.cell(r, 6).value
    g = sheet.cell(r, 7).value
    
    # print raw row
    print(f"Row {r:02d}: B={repr(b)}, C={repr(c)}, D={repr(d)}, E={repr(e)}, F={repr(f)}, G={repr(g)}")

