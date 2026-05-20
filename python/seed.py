from __future__ import annotations

from datetime import date, datetime, time
from pathlib import Path

import openpyxl
import xlrd

from app import create_app
from models import (
    ChecklistItem,
    ChecklistTemplate,
    DailyCheckResult,
    DailyCheckSheet,
    Line,
    ROLE_ADMIN,
    ROLE_LEADER,
    ROLE_MANAGER,
    ROLE_STAFF,
    ROLE_SUPERVISOR,
    User,
    UserLine,
    db,
)


BASE_DIR = Path(__file__).resolve().parent
ACCOUNT_EXCEL = BASE_DIR / "S.d.t.ch.c.xlsx"
TL_EXCEL = BASE_DIR / "TL.DTL.1.xls"
SV_EXCEL = BASE_DIR / "SV.DSV.xls"

ADMIN_TIMES = [time(8, 20), time(9, 20), time(10, 0), time(11, 0), time(13, 0), time(15, 0), time(16, 0)]
SHIFT_TIMES = {
    "Ca 1": [time(6, 0), time(7, 0), time(8, 0), time(9, 0), time(11, 0), time(12, 0), time(13, 0)],
    "Ca 2": [time(14, 0), time(15, 0), time(16, 0), time(17, 0), time(19, 0), time(20, 0), time(21, 0)],
    "Ca 3": [time(22, 0), time(23, 0), time(0, 0), time(1, 0), time(3, 0), time(4, 0), time(5, 0)],
    "Ca hanh chinh": ADMIN_TIMES,
}


def clean(value) -> str:
    return str(value).strip() if value not in {None, ""} else ""


def excel_time(value, fallback: time) -> time:
    if isinstance(value, float):
        total_minutes = round(value * 24 * 60)
        return time((total_minutes // 60) % 24, total_minutes % 60)
    raw = clean(value).splitlines()[0]
    if ":" in raw:
        hour, minute = raw.split(":", 1)
        return time(int(hour), int(minute[:2]))
    return fallback


def normalize_symbol(value) -> str:
    return clean(value).replace("Ｍ", "M").replace("Ｓ", "S").replace("Ｑ", "Q").replace("Ｄ", "D").replace("Ｃ", "C")


def make_user(username: str, full_name: str, role: str, code: str | None = None, password: str = "1") -> User:
    code = code or username
    user = User(
        username=username,
        full_name=full_name,
        employee_code=code,
        role=role,
        department="Production",
        line_name="Ca hanh chinh" if role in {ROLE_MANAGER, ROLE_SUPERVISOR, ROLE_LEADER} else "Ca 1",
        outlook_email=f"{username}@ap.denso.local",
        is_active=True,
    )
    user.set_password(password)
    return user


def unique_key(base: str, used: set[str]) -> str:
    key = base
    index = 2
    while key in used:
        key = f"{base}_{index}"
        index += 1
    used.add(key)
    return key


def parse_account_excel() -> list[dict]:
    wb = openpyxl.load_workbook(ACCOUNT_EXCEL, data_only=True)
    sheet = wb.active
    people = []
    current_mgr = None
    current_sv = None

    names = []
    for row in range(4, sheet.max_row + 1):
        if row == 5:
            continue
        manager_name = clean(sheet.cell(row, 2).value)
        supervisor_name = clean(sheet.cell(row, 3).value)
        leader_name = clean(sheet.cell(row, 4).value)
        if manager_name:
            current_mgr = manager_name
            current_sv = None
            names.append({"role": ROLE_MANAGER, "full_name": manager_name, "parent": None})
        if supervisor_name:
            current_sv = supervisor_name
            names.append({"role": ROLE_SUPERVISOR, "full_name": supervisor_name, "parent": current_mgr})
        if leader_name:
            names.append({"role": ROLE_LEADER, "full_name": leader_name, "parent": current_sv})

    credentials = []
    for row in range(4, sheet.max_row + 1):
        username = clean(sheet.cell(row, 6).value)
        if username:
            credentials.append((username, clean(sheet.cell(row, 7).value) or "1"))

    used_usernames: set[str] = set()
    used_codes: set[str] = set()
    for index, item in enumerate(names):
        username, password = credentials[index] if index < len(credentials) else (f"user{index + 1}", "1")
        username = unique_key(username, used_usernames)
        code = unique_key(username, used_codes)
        people.append({**item, "username": username, "password": password, "code": code})
    return people


def parse_xls_template(path: Path, sheet_index: int, start_row: int, time_col: int, symbol_col: int, content_col: int) -> list[dict]:
    book = xlrd.open_workbook(str(path))
    sheet = book.sheet_by_index(sheet_index)
    rows = []
    current_time = time(8, 0)
    order = 1
    for row_index in range(start_row, sheet.nrows):
        raw_time = sheet.cell_value(row_index, time_col)
        symbol = normalize_symbol(sheet.cell_value(row_index, symbol_col))
        content = clean(sheet.cell_value(row_index, content_col))
        if raw_time not in {"", None}:
            current_time = excel_time(raw_time, current_time)
        if not symbol or not content:
            continue
        if "Nội dung lỗi" in content or "内容" == content:
            break
        rows.append(
            {
                "item_order": order,
                "symbol": symbol,
                "check_time": current_time,
                "time_group": current_time.strftime("%H:%M"),
                "content": content.replace("\n", " ").strip(),
            }
        )
        order += 1
    return rows


def build_template(code: str, name: str, rows: list[dict], lines: list[Line] | None = None) -> ChecklistTemplate:
    template = ChecklistTemplate(template_code=code, template_name=name, version="1.0", is_active=True)
    if not lines:
        for row in rows:
            template.checklist_items.append(
                ChecklistItem(
                    symbol=row["symbol"],
                    check_time=row["check_time"],
                    time_group=row["time_group"],
                    item_order=row["item_order"],
                    category_type=row["symbol"],
                    content=row["content"],
                    content_vi=row["content"],
                    content_en=row["content"],
                    content_ja=row["content"],
                    is_active=True,
                )
            )
        return template

    for line in lines:
        line_times = SHIFT_TIMES[line.line_name]
        for row in rows:
            check_time = line_times[(row["item_order"] - 1) % len(line_times)]
            template.checklist_items.append(
                ChecklistItem(
                    line=line,
                    symbol=row["symbol"],
                    check_time=check_time,
                    time_group=check_time.strftime("%H:%M"),
                    item_order=row["item_order"],
                    category_type=row["symbol"],
                    content=row["content"],
                    content_vi=row["content"],
                    content_en=row["content"],
                    content_ja=row["content"],
                    is_active=True,
                )
            )
    return template


def create_sheet(user: User, template: ChecklistTemplate, target_date: date) -> DailyCheckSheet:
    return DailyCheckSheet(
        user=user,
        template=template,
        check_date=target_date,
        month=target_date.month,
        year=target_date.year,
        line_name=user.line_name,
        department=user.department,
        shift=user.line_name,
        status="draft",
    )


def seed_database() -> None:
    db.drop_all()
    db.create_all()

    lines = [
        Line(line_name="Ca hanh chinh", department="Production", description="Checklist hành chính", is_active=True),
        Line(line_name="Ca 1", department="Production", description="06:00-13:00", is_active=True),
        Line(line_name="Ca 2", department="Production", description="14:00-21:00", is_active=True),
        Line(line_name="Ca 3", department="Production", description="22:00-05:00", is_active=True),
    ]
    db.session.add_all(lines)

    admin = make_user("admin", "System Admin", ROLE_ADMIN, "admin", "1")
    db.session.add(admin)
    db.session.flush()

    users_by_name: dict[str, User] = {}
    leaders: list[User] = []
    for row in parse_account_excel():
        user = make_user(row["username"], row["full_name"], row["role"], row["code"], row["password"])
        if row["role"] == ROLE_SUPERVISOR and row["parent"]:
            user.manager = users_by_name.get(row["parent"])
        elif row["role"] == ROLE_LEADER and row["parent"]:
            user.supervisor = users_by_name.get(row["parent"])
        db.session.add(user)
        db.session.flush()
        users_by_name[row["full_name"]] = user
        if user.role == ROLE_LEADER:
            leaders.append(user)

    for leader in leaders:
        for index in range(1, 4):
            staff_code = f"{leader.employee_code}-SL{index}"
            staff = make_user(
                username=staff_code.lower(),
                full_name=f"{leader.full_name} SL{index}",
                role=ROLE_STAFF,
                code=staff_code,
                password="1",
            )
            staff.leader = leader
            staff.outlook_email = leader.outlook_email
            staff.line_name = f"Ca {index}"
            db.session.add(staff)

    db.session.flush()
    line_by_name = {line.line_name: line for line in lines}
    for user in User.query.filter(User.role != ROLE_ADMIN).all():
        line = line_by_name.get(user.line_name) or line_by_name["Ca hanh chinh"]
        db.session.add(UserLine(user_id=user.id, line_id=line.id))

    tl_rows = parse_xls_template(TL_EXCEL, 0, 8, 0, 1, 2)
    sv_rows = parse_xls_template(SV_EXCEL, 0, 7, 1, 2, 3)
    tl_template = build_template("TL_SL_VN", "Checklist TL/SL", tl_rows, lines)
    sv_template = build_template("SV_DSV_VN", "Checklist SV/DSV", sv_rows)
    db.session.add_all([tl_template, sv_template])
    db.session.flush()

    today = date.today()
    sample_users = User.query.filter(User.role.in_([ROLE_SUPERVISOR, ROLE_LEADER, ROLE_STAFF])).limit(6).all()
    for user in sample_users:
        template = sv_template if user.role == ROLE_SUPERVISOR else tl_template
        sheet = create_sheet(user, template, today)
        db.session.add(sheet)
        db.session.flush()
        items = template.checklist_items
        if user.role != ROLE_SUPERVISOR:
            line = line_by_name.get(user.line_name)
            items = [item for item in items if item.line_id == line.id]
        for item in items:
            db.session.add(
                DailyCheckResult(
                    daily_sheet=sheet,
                    checklist_item=item,
                    user=user,
                    check_date=today,
                    symbol=item.symbol,
                    check_time=item.check_time,
                    content=item.content_vi,
                    result="",
                )
            )

    db.session.commit()


app = create_app()


if __name__ == "__main__":
    with app.app_context():
        seed_database()
    print("Seed completed.")
    print("Accounts: admin / 1; Excel accounts use password 1.")
